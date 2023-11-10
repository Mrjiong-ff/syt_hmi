#!coding=utf-8
import argparse
import json

import numpy as np
import openpyxl

parser = argparse.ArgumentParser("解析参数表格，自动生成参数配置前端所需的json文件")
parser.add_argument("--input", default="", help="输入的xlsx文件路径，请注意检查文件是否被加密")


class SYTParam(object):
    def __init__(self):
        self.dtype = "float32"
        self.length = "1"
        self.min = "0"
        self.max = "0"
        self.value = ""
        self.value_list = []
        self.storable = "1"
        self.comment = ""
        self.modify_freq = 0

    def to_dict(self):
        result = {}
        if len(self.value_list) > 0:
            result = {
                "dtype": self.dtype,
                "length": self.length,
                "min": self.min,
                "max": self.max,
                "value": self.value_list,
                "storable": self.storable,
                "comment": self.comment,
                "modify_freq": self.modify_freq,
            }
        else:
            result = {
                "dtype": self.dtype,
                "length": self.length,
                "min": self.min,
                "max": self.max,
                "value": self.value,
                "storable": self.storable,
                "comment": self.comment,
                "modify_freq": self.modify_freq,
            }
        if result["dtype"] == "bool":
            result["dtype"] = "uchar"
        return result


class SYTParamSheet(object):
    from openpyxl.worksheet.worksheet import Worksheet as xl_worksheet

    def __init__(self, sheet: xl_worksheet):
        self.sheet_name = sheet.title
        self.signed_dtype_list = [
            "float32",
            "int32",
        ]
        self.param_dict = {}
        for i in range(1, sheet.max_row):
            row_idx = i + 1
            param_element = SYTParam()
            if sheet.cell(row_idx, 1).value is None:
                continue
            param_element.dtype = self.parse_dtype(sheet.cell(row_idx, 3).value)
            if param_element.dtype == "bool":
                param_element.min = "0"
                param_element.max = "1"
            else:
                param_element.min, param_element.max = self.parse_range(
                    sheet.cell(row_idx, 5).value, param_element.dtype
                )
            value, param_element.length = self.parse_value(
                sheet.cell(row_idx, 2).value, sheet.cell(row_idx, 4).value
            )
            if isinstance(value, list):
                param_element.value_list = value
            else:
                param_element.value = value
            param_element.comment = self.parse_comment(sheet.cell(row_idx, 8).value)
            param_element.modify_freq = self.parse_modify_freq(
                sheet.cell(row_idx, 7).value
            )
            self.param_dict[str(sheet.cell(row_idx, 1).value)] = param_element.to_dict()

    def parse_dtype(self, type_value):
        if type_value == "float":
            return "float32"
        if type_value == "int8":
            return "char"
        if type_value == "uint8":
            return "uchar"
        if type_value == "double":
            return "float64"
        if type_value == "int":
            return "int32"
        return str(type_value)

    def parse_range(self, range_value, dtype):
        if range_value is None:
            return str(self.get_bottom_limit(dtype)), str(self.get_top_limit(dtype))
        limit_tuple = str(range_value).split("~")
        if limit_tuple[0] == "":
            limit_tuple[0] = str(self.get_bottom_limit(dtype))
        if limit_tuple[1] == "":
            limit_tuple[1] = str(self.get_top_limit(dtype))
        return limit_tuple

    def get_bottom_limit(self, dtype):
        if dtype == "float32":
            return np.finfo(np.float32).min
        if dtype == "float64":
            return np.finfo(np.float64).min
        if dtype == "int32":
            return np.iinfo(np.int32).min
        if dtype == "uint32":
            return np.iinfo(np.uint32).min
        if dtype == "int64":
            return np.iinfo(np.int64).min
        if dtype == "uint64":
            return np.iinfo(np.uint64).min
        if dtype == "char":
            return np.iinfo(np.int8).min
        if dtype == "uchar":
            return np.iinfo(np.uint8).min
        return 0

    def get_top_limit(self, dtype):
        if dtype == "float32":
            return np.finfo(np.float32).max
        if dtype == "float64":
            return np.finfo(np.float64).max
        if dtype == "int32":
            return np.iinfo(np.int32).max
        if dtype == "uint32":
            return np.iinfo(np.uint32).max
        if dtype == "int64":
            return np.iinfo(np.int64).max
        if dtype == "uint64":
            return np.iinfo(np.uint64).max
        if dtype == "char":
            return np.iinfo(np.int8).max
        if dtype == "uchar":
            return np.iinfo(np.uint8).max
        return 0

    def parse_value(self, default_value, value_dimension):
        # 首先解析值的维度，是单个数值还是某个长度的数组
        value_dimension_str = str(value_dimension)
        if value_dimension_str[0] == "[":
            value_dimension_str = value_dimension_str[1:]
        if value_dimension_str[-1] == "]":
            value_dimension_str = value_dimension_str[:-1]
        try:
            dimension = int(value_dimension_str)
        except Exception as e:
            print(e)
            print("参数维度数据格式错误")
            exit(-2)
        if default_value is None:
            return "0", str(dimension)
        default_value_str = str(default_value)
        if default_value_str[0] == "[":
            default_value_str = default_value_str[1:]
        if default_value_str[-1] == "]":
            default_value_str = default_value_str[:-1]
        if default_value_str.find(",") >= 0:
            default_values = default_value_str.split(",")
            if len(default_values) != dimension:
                print("数组类型默认参数格式错误！请填写单个数字或dimension个数字并以逗号隔开")
                exit(-1)
            return default_values, str(dimension)
        else:
            value_str = ""
            if dimension > 1:
                value_list = [default_value_str] * dimension
                return value_list, str(dimension)
            else:
                value_str = str(default_value_str)
                return str(value_str), str(dimension)

    def parse_modify_freq(self, modify_freq_value):
        modify_freq_str = str(modify_freq_value)
        if modify_freq_str == "仅出厂":
            return 0
        if modify_freq_str == "低":
            return 1
        if modify_freq_str == "高":
            return 2
        return -1

    def parse_comment(self, comment_value):
        return str(comment_value)

    def to_dict(self, freq_level=2):
        result = {}
        for key in self.param_dict:
            if self.param_dict[key]["modify_freq"] == freq_level:
                result[key] = self.param_dict[key]
        return result


if __name__ == "__main__":
    args = parser.parse_args()
    wb = openpyxl.load_workbook(args.input)
    total_param_sheets = {}
    for sheet in wb.worksheets:
        param_sheet = SYTParamSheet(sheet)
        total_param_sheets[param_sheet.sheet_name] = param_sheet
    high_freq_dict = {}
    low_freq_dict = {}
    only_factory_dict = {}
    for sheet_name in total_param_sheets:
        high_freq_dict[sheet_name] = total_param_sheets[sheet_name].to_dict(2)
        low_freq_dict[sheet_name] = total_param_sheets[sheet_name].to_dict(1)
        only_factory_dict[sheet_name] = total_param_sheets[sheet_name].to_dict(0)

    high_freq_param_file = open("param.json", "w")
    low_freq_param_file = open("param_low_freq.json", "w")
    only_factory_param_file = open("param_factory_only.json", "w")

    json.dump(high_freq_dict, high_freq_param_file, indent=4, ensure_ascii=False)
    json.dump(low_freq_dict, low_freq_param_file, indent=4, ensure_ascii=False)
    json.dump(only_factory_dict, only_factory_param_file, indent=4, ensure_ascii=False)
    high_freq_param_file.close()
    low_freq_param_file.close()
    only_factory_param_file.close()
